import re
import csv
from pathlib import Path
from openpyxl import load_workbook
from datetime import timedelta

def clean_teacher(name: str) -> str:
    """Очищает имя преподавателя, сохраняя инициалы"""
    if not name:
        return ''
    # Убираем дублирование "Сокольников В. В.Сокольников В. В."
    parts = re.split(r'(?<=\.)\s*(?=[А-Я])', name)
    if len(parts) > 1 and parts[0] == parts[1]:
        return parts[0]
    if len(name) % 2 == 0 and name[:len(name)//2] == name[len(name)//2:]:
        return name[:len(name)//2]
    # Убираем лишние пробелы
    name = re.sub(r'\s+', ' ', name)
    name = re.sub(r'\.\.+', '.', name)
    # Убираем цифры и слово "нед"
    name = re.sub(r'\d+\s*нед\.', '', name)
    name = re.sub(r'[\d,]+', '', name)
    name = name.strip()
    # Если имя пустое после очистки
    if not name:
        return ''
    return name

def parse_time_from_text(text: str):
    pattern = r'с\s*(\d{1,2}:\d{2})\s*-до\s*(\d{1,2}:\d{2})'
    m = re.search(pattern, text)
    if m:
        return m.group(1), m.group(2)
    return None, None

def parse_weeks_from_text(text: str):
    m = re.search(r'([\d,]+)\s*нед\.', text)
    return m.group(1).strip() if m else ''

def is_auditorium(text: str) -> bool:
    if not text:
        return False
    text = str(text).strip()
    # Шаблоны: 402/5, 310/2, 1411, 1407, 6343, 218/1, 327/1, 406/5, 301/2
    return bool(re.match(r'^[А-Я]?\d{1,4}([/\-]\d+)?$', text))

def remove_seconds(time_str: str) -> str:
    if not time_str:
        return ''
    parts = time_str.split(':')
    return f"{parts[0]}:{parts[1]}" if len(parts) >= 2 else time_str

def parse_lesson_cell(cell_value: str):
    if not cell_value or str(cell_value).strip() == '':
        return None
    
    # Если ячейка - только аудитория, возвращаем None
    if is_auditorium(str(cell_value).strip()):
        return None
    
    lines = str(cell_value).strip().split('\n')
    lines = [l.strip() for l in lines if l.strip()]
    if not lines:
        return None

    start_time, end_time = parse_time_from_text(lines[0])
    if start_time:
        lines.pop(0)

    lesson_type = ''
    subject = ''
    teacher = ''
    weeks = ''

    type_keywords = ['лекционные', 'лабораторные', 'практические']
    type_idx = -1
    for i, line in enumerate(lines):
        if any(kw in line.lower() for kw in type_keywords):
            lesson_type = line.strip()
            type_idx = i
            break
    if type_idx == -1 and lines:
        lesson_type = lines[0]
        type_idx = 0

    if type_idx + 1 < len(lines):
        subject = lines[type_idx + 1].strip()
    if type_idx + 2 < len(lines):
        teacher_line = lines[type_idx + 2].strip()
        teacher_line = re.sub(r'^\(|\)$', '', teacher_line)
        weeks = parse_weeks_from_text(teacher_line)
        teacher_line = re.sub(r'[\d,]+\s*нед\.', '', teacher_line).strip()
        teacher = clean_teacher(teacher_line)
    elif type_idx + 1 < len(lines) and not teacher:
        combined = lines[type_idx + 1].strip()
        match = re.match(r'^(.*?)\s*\((.*?)\)$', combined)
        if match:
            subject, teacher = match.groups()
            teacher = clean_teacher(teacher)
            weeks = parse_weeks_from_text(teacher) or parse_weeks_from_text(combined)
            teacher = re.sub(r'[\d,]+\s*нед\.', '', teacher).strip()
        else:
            subject = combined

    if start_time:
        start_time = remove_seconds(start_time)
    if end_time:
        end_time = remove_seconds(end_time)

    return {
        'type': lesson_type,
        'subject': subject,
        'teacher': teacher,
        'start_time': start_time,
        'end_time': end_time,
        'weeks': weeks
    }

def find_auditorium_column(ws):
    """Ищет столбец с названием 'Ауд'"""
    for row in range(1, min(20, ws.max_row + 1)):
        for col in range(1, min(25, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if val and isinstance(val, str) and re.search(r'Ауд', val, re.IGNORECASE):
                return col
    return None

def find_subgroup_header(ws):
    """Находит строку с подписями подгрупп"""
    for row in range(1, min(25, ws.max_row + 1)):
        subgroup_cols = []
        for col in range(1, min(25, ws.max_column + 1)):
            val = ws.cell(row, col).value
            if val and isinstance(val, str) and re.search(r'\d+\s*п/?г', val, re.IGNORECASE):
                subgroup_cols.append(col)
        if subgroup_cols:
            return row, subgroup_cols
    return None, None

def extract_timetable(excel_path, output_csv, group_name):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    common_aud_col = find_auditorium_column(ws)
    header_row, subgroup_cols = find_subgroup_header(ws)
    
    if subgroup_cols:
        subgroups = []
        for idx, col in enumerate(subgroup_cols, 1):
            subgroups.append({
                'subgroup_num': idx,
                'lesson_col': col,
                'aud_col': col + 1 if col + 1 <= ws.max_column else None
            })
        print(f"Обнаружено {len(subgroups)} подгрупп, общий столбец аудиторий: {common_aud_col}")
    else:
        subgroups = [{'subgroup_num': 1, 'lesson_col': 4, 'aud_col': None}]
        print(f"Подгруппы не найдены. Используется 1 подгруппа, столбец аудиторий: {common_aud_col}")

    start_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == 'Понедельник':
            start_row = row
            break
    if not start_row:
        raise ValueError("Не найдена строка с 'Понедельник'")

    records = []
    days_order = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
    current_day = None
    prev_time_slot = None
    # Кэш для наследования аудиторий
    last_aud_for_slot = {sg['subgroup_num']: '' for sg in subgroups}
    slot_cache = {}  # (day, time_slot) -> аудитория для подгруппы

    for row in range(start_row, ws.max_row + 1):
        day_cell = ws.cell(row, 1).value
        time_slot_cell = ws.cell(row, 2).value

        if day_cell in days_order:
            current_day = day_cell
            prev_time_slot = None
            last_aud_for_slot = {sg['subgroup_num']: '' for sg in subgroups}
            slot_cache = {}
        if not current_day:
            continue

        current_time_slot = str(time_slot_cell).strip() if time_slot_cell else ''
        slot_key = (current_day, current_time_slot)
        is_new_time_slot = (current_time_slot and current_time_slot != prev_time_slot)
        if is_new_time_slot:
            last_aud_for_slot = {sg['subgroup_num']: '' for sg in subgroups}
        prev_time_slot = current_time_slot

        # Общая аудитория для строки
        common_aud = ''
        if common_aud_col:
            aud_val = ws.cell(row, common_aud_col).value
            if aud_val and str(aud_val).strip():
                common_aud = str(aud_val).strip()

        for sg in subgroups:
            lesson_col = sg['lesson_col']
            aud_col = sg['aud_col']
            subgroup_num = sg['subgroup_num']
            
            lesson_cell = ws.cell(row, lesson_col).value
            if not lesson_cell:
                continue

            # Если ячейка содержит только аудиторию - сохраняем в кэш
            if is_auditorium(str(lesson_cell).strip()):
                last_aud_for_slot[subgroup_num] = str(lesson_cell).strip()
                slot_cache[slot_key] = slot_cache.get(slot_key, {})
                slot_cache[slot_key][subgroup_num] = str(lesson_cell).strip()
                continue

            parsed = parse_lesson_cell(lesson_cell)
            if not parsed or not parsed['subject']:
                continue

            # Определяем аудиторию по приоритету:
            # 1. Индивидуальная аудитория подгруппы
            # 2. Аудитория из кэша для этого таймслота
            # 3. Общая аудитория строки
            aud = ''
            if aud_col:
                aud_val = ws.cell(row, aud_col).value
                if aud_val and str(aud_val).strip():
                    aud = str(aud_val).strip()
            if not aud and last_aud_for_slot.get(subgroup_num):
                aud = last_aud_for_slot[subgroup_num]
            if not aud and common_aud:
                aud = common_aud

            # Если нашли аудиторию - сохраняем в кэш для этой подгруппы
            if aud:
                last_aud_for_slot[subgroup_num] = aud
                if slot_key not in slot_cache:
                    slot_cache[slot_key] = {}
                slot_cache[slot_key][subgroup_num] = aud

            # Если аудитории нет, но есть у другой подгруппы в этом таймслоте - наследуем
            if not aud and slot_key in slot_cache:
                for other_subgroup, other_aud in slot_cache[slot_key].items():
                    if other_aud:
                        aud = other_aud
                        break

            # Время
            start_time = parsed['start_time']
            end_time = parsed['end_time']
            if not start_time and is_new_time_slot and time_slot_cell:
                parts = str(time_slot_cell).split(' - ')
                start_time = remove_seconds(parts[0]) if parts else ''
                end_time = remove_seconds(parts[1]) if len(parts) > 1 else ''
            if not start_time:
                continue
            if not end_time:
                try:
                    dt = timedelta(hours=int(start_time[:2]), minutes=int(start_time[3:]))
                    dt += timedelta(hours=1, minutes=30)
                    end_time = f"{dt.hours:02d}:{dt.minutes:02d}"
                except:
                    end_time = ''

            records.append({
                'day': current_day,
                'time_slot': time_slot_cell or '',
                'start_time': start_time,
                'end_time': end_time,
                'type': parsed['type'],
                'subject': parsed['subject'],
                'teacher': parsed['teacher'],
                'auditorium': aud,
                'subgroup': subgroup_num,
                'group_name': group_name,
                'weeks': parsed['weeks']
            })

    # Удаление дубликатов
    unique = {}
    for rec in records:
        key = (rec['day'], rec['start_time'], rec['subject'], rec['subgroup'])
        if key not in unique:
            unique[key] = rec
    unique_records = list(unique.values())

    with open(output_csv, 'w', newline='', encoding='utf-8-sig') as f:
        fieldnames = ['day', 'time_slot', 'start_time', 'end_time', 'type', 'subject',
                      'teacher', 'auditorium', 'subgroup', 'group_name', 'weeks']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(unique_records)

    print(f"Группа {group_name}: обработано {len(unique_records)} занятий. CSV: {output_csv}")

def normalize_group_name(filename: str) -> str:
    name = Path(filename).stem
    name = re.sub(r'\s*\(\d+\)\s*$', '', name)
    trans = str.maketrans({
        'b': 'б', 'B': 'Б', 'k': 'к', 'K': 'К', 'p': 'п', 'P': 'П',
        'o': 'о', 'O': 'О', 'c': 'с', 'C': 'С', 'e': 'е', 'E': 'Е',
        'a': 'а', 'A': 'А', 't': 'т', 'T': 'Т', 'i': 'и', 'I': 'И',
        'y': 'ы', 'Y': 'Ы', 'x': 'кс', 'X': 'Кс'
    })
    name = name.translate(trans)
    name = re.sub(r'(?i)po', 'по', name)
    name = name.replace('_', '-')
    return name.lower()

if __name__ == '__main__':
    script_dir = Path(__file__).parent
    xlsx_folder = script_dir / 'xlsx_files'
    output_folder = script_dir / 'csv_output'
    output_folder.mkdir(exist_ok=True)

    if not xlsx_folder.exists():
        print(f"Папка {xlsx_folder} не найдена. Создайте её и положите туда XLSX-файлы.")
        exit(1)

    xlsx_files = list(xlsx_folder.glob('*.xlsx'))
    if not xlsx_files:
        print(f"В папке {xlsx_folder} нет файлов .xlsx")
        exit(0)

    for xlsx_path in xlsx_files:
        group_name = normalize_group_name(xlsx_path.name)
        csv_path = output_folder / f"{group_name}.csv"
        print(f"\nОбработка: {xlsx_path.name} -> группа {group_name}")
        extract_timetable(str(xlsx_path), str(csv_path), group_name)

    print(f"\nВсе CSV сохранены в {output_folder}")